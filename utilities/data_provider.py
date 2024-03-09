import openpyxl
from pathlib import Path
from typing import IO, Iterator, Dict, Any
import subprocess, tempfile, csv
from python_calamine import CalamineWorkbook

parent_path = Path(__file__).resolve().parents[1]
routines_file = f'{parent_path}/test_data/Routines.xlsx'
data_file = f'{parent_path}/test_data/testdata.xlsx'


def get_data(sheet_name):
    try:
        test_data = openpyxl.load_workbook(data_file)
        sheet = test_data[sheet_name]
        total_rows = sheet.max_row
        total_cols = sheet.max_column
        print("total cols are ", str(total_cols))
        print("total rows are ", str(total_rows))
        main_list = []

        for i in range(2, total_rows + 1):
            dataList = []
            for j in range(1, total_cols + 1):
                data = sheet.cell(row=i, column=j).value
                dataList.insert(j, data)
            main_list.insert(i, dataList)
            print(main_list)
        return main_list
    except PermissionError:
        print('Permission error')


def get_authentication(account_type):
    try:
        workbook = openpyxl.load_workbook(data_file)
        sheet = workbook["authentication"]
        total_rows = sheet.max_row
        total_cols = sheet.max_column
        print("total cols are ", str(total_cols))
        print("total rows are ", str(total_rows))
        main_data = []

        for i in range(2, total_rows + 1):
            temp_data = []
            for j in range(1, total_cols + 1):
                data = sheet.cell(row=i, column=j).value
                if str(data).lower() == str(account_type).lower():
                    temp_data.insert(j, sheet.cell(row=i, column=j + 1).value)
                    temp_data.insert(j, sheet.cell(row=i, column=j + 2).value)
                    main_data.insert(i, temp_data)
                    return main_data
    except PermissionError:
        print('Permission error')


def get_routine_from_routine_management():
    try:
        workbook = openpyxl.load_workbook(routines_file)
        sheet = workbook['EDITED']
        total_rows = sheet.max_row
        total_cols = sheet.max_column
        print("total cols are ", str(total_cols))
        print("total rows are ", str(total_rows))
        main_data = []
        for i in range(1, total_rows + 1):
            temp_data = []
            for j in range(1, 39):
                data = sheet.cell(row=i, column=j).value
                temp_data.insert(j, sheet.cell(row=i, column=j + 1).value)
                temp_data.insert(j, sheet.cell(row=i, column=j + 2).value)
                main_data.insert(i, temp_data)
        yield main_data

    except PermissionError:
        print('Permission error')


def iter_excel_openpyxl_without_build_in(file: IO[bytes]) -> Iterator[dict[str, object]]:
    workbook = openpyxl.load_workbook(file, read_only=True)
    sheet = workbook['EDITED']
    total_rows = sheet.max_row
    # total_cols = sheet.max_column
    check_lists = {}
    for row in range(1, total_rows):
        check_list_item = {}
        for col in range(0, 39):
            title = sheet.cell(row=1, column=col + 1).value
            cell = sheet.cell(row=row + 1, column=col + 1).value
            check_list_item[title] = cell
            check_lists[row] = check_list_item
    print(f'check_list: {check_lists}')
    print("stop")


def iter_excel_openpyxl(file: IO[bytes]) -> Iterator[dict[Any, Any]]:
    workbook = openpyxl.load_workbook(routines_file, read_only=True)
    rows = workbook.active.rows
    headers = [str(cell.value) for cell in next(rows)]
    for row in rows:
        print(f'{row}')
        yield dict(zip(headers, (cell.value for cell in row)))


def iter_excel_calamine(file: IO[bytes]) -> Iterator[dict[str, object]]:
    workbook = CalamineWorkbook.from_path(file)
    sheet = workbook.get_sheet_by_name('EDITED').to_python(skip_empty_area=True)
    headers = sheet[0]
    for row in sheet:
        yield dict(zip(headers, row))


def collect_usable_record_from_routine_file(file):
    with open(file, 'rb') as f:
        rows = iter_excel_calamine(file)
        routines = []
        for row in rows:
            if not str(row['Checklist Category']) == 'NA' and not str(row['Checklist Category']) == '':
                routines.append(row)

        print(routines)


if __name__ == '__main__':
    collect_usable_record_from_routine_file(routines_file)
